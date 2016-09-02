from django import forms
from django.db import transaction
from django.forms import ModelForm
from django.forms.fields import TypedChoiceField
from openpyxl import load_workbook

from .models import Cattle
from user.models import User
from user.forms import UserForm


class CattleForm(ModelForm):
    producer = forms.ModelChoiceField(queryset=User.objects.all(), required=False)

    class Meta:
        model = Cattle
        fields = ['lot_number', 'breed',
                  'sex', 'female_type',
                  'registration_number', 'bull_name',
                  'sire', 'dame',
                  'scrotal_circumference', 'birth_weight',
                  'weaning_weight', 'yearling_weight',
                  'residual_average_daily_gain', 'heifer_pregnancy',
                  'calving_ease_maternal', 'maternal_milk',
                  'mature_weight', 'mature_height',
                  'cow_energy_value', 'carcass_weight',
                  'marbling', 'ribeye_area',
                  'fat_thickness', 'video_url']


class ImportCattleForm(forms.Form):
    cattle_xlsx = forms.FileField(label='Cattle Excel Spreadshet (.xlsx)')

    def clean_cattle_xlsx(self):
        cattle_xlsx = self.cleaned_data['cattle_xlsx']
        try:
            load_workbook(self.cleaned_data['cattle_xlsx'])
        except Exception:
            raise forms.ValidationError('Failed to load spreadsheet!')
        return cattle_xlsx

    @transaction.atomic
    def import_cattle(self):
        """
        Imports an Excel spreadsheet of bulls, with checks for new producers.

        Returns a tuple (is_valid, form) where is_valid is a boolean indicating
        if any form errors occured and the corresponding form instance.
        """
        # TODO: Break this out into multiple functions
        workbook = load_workbook(self.cleaned_data['cattle_xlsx'])
        for row_num, row in enumerate(workbook.active):
            # Skip column header and empty rows
            if row[0].value is None or 'Producer' in row[0].value:
                continue
            producer = {k: v for k, v in zip(
                ('name', 'email', 'address', 'address2', 'city', 'state', 'zip'),
                [cell.value for cell in row[0:7]])}
            try:
                producer = User.objects.get(email=producer['email'])
            except User.DoesNotExist:
                user_form = UserForm(producer)
                if not user_form.is_valid():
                    user_form.add_error(None, 'Error with producer on row {}!'.format(row_num + 1))
                    return False, user_form
                producer = user_form.save(commit=False)
                producer.type = User.PRODUCER
            bull = {k: v for k, v in zip(
                [field for field in CattleForm.Meta.fields],
                [cell.value for cell in row[7:]])}
            if not Cattle.objects.filter(producer=producer, bull_name=bull['bull_name']).exists():
                # Convert verbose choices in spreadsheet to shortened versions
                # TODO: This isn't safe - will easily break if any choice fields change.
                choice_fields = {field: field_type for field, field_type
                                 in CattleForm.base_fields.items()
                                 if type(field_type) is TypedChoiceField}
                for field, field_type in choice_fields.items():
                    bull[field] = bull[field].replace(' ', '_').lower()
                cattle_form = CattleForm(bull)
                if not cattle_form.is_valid():
                    cattle_form.add_error(None, 'Error with bull on row {}!'.format(row_num + 1))
                    return False, cattle_form
                producer.save()
                cattle_form.instance.producer_id = producer.id
                cattle_form.save()
        return True, {}
